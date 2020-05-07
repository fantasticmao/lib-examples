#!/usr/local/bin/python3
# Python 练习4 - Excel

# doc http://openpyxl.readthedocs.io/
from openpyxl import Workbook
from openpyxl import load_workbook

file_path = "../../resources/test.xlsx"


# 创建工作簿
def create_workbook():
    workbook = Workbook()
    # 修改当前激活的工作表标题
    workbook.active.title = "test sheet".title()
    return workbook


# 访问工作表
def access_worksheet(workbook):
    # 获取当前激活的工作表标题 A1:C3 的单元格
    for row in workbook.active["A1":"C3"]:
        for cell in row:
            # 设置单元格的内容
            cell.value = "hello " + cell.coordinate


# 保存工作薄
def save_workbook(workbook):
    workbook.save(filename=file_path)


# 打印工作表
def print_worksheet(workbook):
    for row in workbook.active.values:
        for value in row:
            print(value)


def main():
    workbook = create_workbook()
    access_worksheet(workbook)
    save_workbook(workbook)

    # 读取工作薄
    workbook = load_workbook(filename=file_path)
    print_worksheet(workbook)


if __name__ == "__main__":
    main()
